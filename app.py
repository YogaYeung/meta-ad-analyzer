"""
META广告竞品拆解分析工具
专为跨境独立站运营者设计
分析竞品广告素材，输出20维度结构化报告 + 可迁移创意公式 + 品牌落地建议
"""

import streamlit as st
from google import genai
from google.genai import types
import pandas as pd
import json
import time
import io
import os
import tempfile
from pathlib import Path

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# ═══════════════════════════════════════════════════════
# 页面配置
# ═══════════════════════════════════════════════════════
st.set_page_config(
    page_title="META广告竞品分析神器",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ═══════════════════════════════════════════════════════
# 核心 Prompt —— 单条素材分析（20个维度）
# ═══════════════════════════════════════════════════════
SINGLE_AD_PROMPT = """你是一名专业的Meta广告创意策略分析师，专注于跨境电商DTC品牌广告研究。

请深度分析这个广告素材的视觉内容（画面、文字、音频、节奏、风格等所有可见信息），严格按JSON格式输出，所有字段必须填写，不确定的填"无法判断"。

只输出JSON，不要任何其他文字：
{{
  "素材编号": "{ad_id}",
  "素材类型": "图片/视频/GIF/轮播图（根据素材实际类型填写）",
  "素材形式": "UGC真实用户/专业棚拍/混合剪辑/动画/屏幕录制/开箱测评/对比展示/教程演示/其他",
  "广告感判断": "强广告感/中等广告感/弱广告感（原生内容感）",
  "素材时长": "秒数，图片填0，视频填实际秒数",
  "是否前三秒出现产品": "是/否/不适用（图片类）",
  "Hook类型": "痛点刺激型/好奇悬念型/冲突反差型/数据权威型/场景代入型/利益诱惑型/情绪共鸣型/直接展示型",
  "Hook具体表达方式": "视频前3秒画面+音频内容，或图片主视觉+主标题的具体描述",
  "Hook对应Persona": "这个Hook主要吸引什么样的人（年龄、性别、当前状态、核心痛点）",
  "Hook抗划走原因": "从心理学角度分析这个Hook为何能阻止用户划走，背后的心理机制是什么",
  "主要Persona": "完整用户画像：年龄段、性别、生活状态、核心痛点、消费动机",
  "投放意图": "冷启动拉新/兴趣再营销/加购放弃再营销/老客唤回/品牌建设",
  "是否排除某类人群": "是（描述被排除的人群特征）/否/无法判断",
  "核心信息主张": "一句话概括这个广告的核心Messaging",
  "主张类型": "功能利益型/情感共鸣型/社会认同型/价格促销型/稀缺紧迫型/生活方式型",
  "本质卖点": "去掉所有表达包装后，这个产品/品牌最底层的核心价值是什么",
  "是否具备问题证明结果结构": "是（完整三段式：问题→证明→结果）/部分（缺少：X段）/否",
  "是否包含社会证明": "是（类型：用户评论/销量数据/KOL推荐/媒体背书/专业认证）/否",
  "CTA类型": "Shop Now/Learn More/Get Offer/Sign Up/Download/Watch More/无明显CTA",
  "适合投放阶段": "TOF冷启动/MOF考虑层/BOF转化层/全阶段适用"
}}"""

# ═══════════════════════════════════════════════════════
# 核心 Prompt —— 综合报告（可迁移创意公式 + 落地建议）
# ═══════════════════════════════════════════════════════
SUMMARY_PROMPT = """你是一名专业的Meta广告创意策略分析师，帮助跨境DTC品牌从竞品广告中系统性地提取可用洞察。

以下是对{n}条竞品广告的逐条分析结果：
{analysis_json}

---
我的品牌信息：
品牌名称：{brand_name}
竞品来源品牌：{competitor_name}
我的产品：{brand_product}
我的目标人群：{brand_persona}
我的核心卖点：{brand_usp}
---

请输出完整的策略分析报告（中文，Markdown格式）：

## 一、竞品广告整体策略洞察

### 素材形式偏好分析
（竞品主要用哪些形式？比例如何？背后逻辑是什么？）

### 高频Hook策略规律
（最常用的Hook类型是什么？有没有重复出现的开场模式？）

### 核心Persona画像共性
（竞品广告面向的主要目标人群特征）

### 漏斗阶段布局分析
（TOF/MOF/BOF各占多少？整体投放策略倾向？）

### 竞品核心Messaging规律
（竞品反复主张的核心信息是什么？说明他们认为最有效的卖点是？）

---

## 二、可迁移创意公式（3-5个）

针对每个公式，按以下格式输出：

**公式X：[简洁公式名称]**
- 📐 结构模板：[Hook类型] + [内容节奏] + [结尾CTA]
- 🎯 适用场景：
- ✅ 核心成功要素：
- 💬 参考Hook示例：

（重复以上格式输出每个公式）

---

## 三、针对[{brand_name}]的落地建议

### 优先测试的素材形式（按优先级从高到低）

### 推荐Hook策略（附具体文案示例，基于我品牌的卖点改写）

### Messaging框架建议
（基于竞品规律，我应该主打哪个方向？具体怎么说？）

### TOF / MOF / BOF 各阶段打法建议

### 需要规避的创意误区
（看完竞品，哪些常见做法实际效果可能不好？）"""

# ═══════════════════════════════════════════════════════
# 20列的标准输出顺序
# ═══════════════════════════════════════════════════════
COLUMNS = [
    "素材编号", "素材类型", "素材形式", "广告感判断",
    "素材时长", "是否前三秒出现产品",
    "Hook类型", "Hook具体表达方式", "Hook对应Persona", "Hook抗划走原因",
    "主要Persona", "投放意图", "是否排除某类人群",
    "核心信息主张", "主张类型", "本质卖点",
    "是否具备问题证明结果结构", "是否包含社会证明",
    "CTA类型", "适合投放阶段"
]

# ═══════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════

def parse_json_response(text: str) -> dict:
    """从Gemini响应中稳健地提取JSON"""
    text = text.strip()
    # 去掉 markdown 代码块
    for marker in ["```json", "```JSON", "```"]:
        if marker in text:
            parts = text.split(marker)
            for part in parts[1:]:
                candidate = part.split("```")[0].strip()
                try:
                    return json.loads(candidate)
                except Exception:
                    continue
    # 直接尝试解析
    return json.loads(text)


def analyze_single_ad(client, model_name: str, ad_id: str,
                      file_bytes: bytes = None, file_ext: str = None) -> dict:
    """调用Gemini分析一条广告素材"""
    prompt = SINGLE_AD_PROMPT.format(ad_id=ad_id)
    content_parts = [prompt]
    tmp_path = None

    try:
        if file_bytes and file_ext:
            ext = file_ext.lower()

            # 图片：直接传 PIL Image
            if ext in {'.jpg', '.jpeg', '.png', '.gif', '.webp'}:
                if PIL_AVAILABLE:
                    img = Image.open(io.BytesIO(file_bytes))
                    content_parts.append(img)

            # 视频：通过 Files API 上传
            elif ext in {'.mp4', '.mov', '.avi', '.webm'}:
                with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                    tmp.write(file_bytes)
                    tmp_path = tmp.name
                try:
                    vf = client.files.upload(path=tmp_path)
                    # 等待处理完成（最多60秒）
                    for _ in range(20):
                        if vf.state.name != "PROCESSING":
                            break
                        time.sleep(3)
                        vf = client.files.get(name=vf.name)
                    if vf.state.name == "ACTIVE":
                        content_parts.append(vf)
                    else:
                        st.warning("⚠️ 视频处理超时，将仅分析画面截图")
                except Exception as ve:
                    st.warning(f"⚠️ 视频上传遇到问题（{str(ve)[:80]}）")

        resp = client.models.generate_content(
            model=model_name,
            contents=content_parts
        )
        return parse_json_response(resp.text)

    except json.JSONDecodeError:
        raw = resp.text if 'resp' in locals() else "无响应"
        row = {col: "解析失败" for col in COLUMNS}
        row["素材编号"] = ad_id
        row["核心信息主张"] = f"JSON解析失败：{raw[:200]}"
        return row

    except Exception as e:
        row = {col: "分析错误" for col in COLUMNS}
        row["素材编号"] = ad_id
        row["核心信息主张"] = str(e)[:200]
        return row

    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def build_excel(df: pd.DataFrame, summary: str) -> bytes:
    """生成格式化的Excel报告"""
    from openpyxl.styles import PatternFill, Font, Alignment

    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:

        # Sheet 1：素材分析明细
        df.to_excel(writer, sheet_name='素材分析明细', index=False)
        ws = writer.sheets['素材分析明细']

        # 表头样式
        hfill = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 32

        # 数据行样式
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 72
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 列宽
        col_widths = {
            "素材编号": 10, "素材类型": 10, "素材形式": 16,
            "广告感判断": 14, "素材时长": 10, "是否前三秒出现产品": 18,
            "Hook类型": 15, "Hook具体表达方式": 38, "Hook对应Persona": 32,
            "Hook抗划走原因": 38, "主要Persona": 38, "投放意图": 18,
            "是否排除某类人群": 22, "核心信息主张": 38, "主张类型": 15,
            "本质卖点": 32, "是否具备问题证明结果结构": 24,
            "是否包含社会证明": 22, "CTA类型": 14, "适合投放阶段": 16
        }
        for i, col_name in enumerate(df.columns, 1):
            letter = ws.cell(1, i).column_letter
            ws.column_dimensions[letter].width = col_widths.get(col_name, 20)

        # Sheet 2：综合策略报告
        ws2 = writer.book.create_sheet("综合策略报告")
        ws2.column_dimensions['A'].width = 120
        ws2['A1'] = summary
        ws2['A1'].alignment = Alignment(wrap_text=True, vertical='top')
        ws2.row_dimensions[1].height = 800

    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════
# 主应用
# ═══════════════════════════════════════════════════════

def main():

    # ── 顶部标题 ──
    st.title("🎯 META广告竞品拆解分析工具")
    st.markdown(
        "**专为跨境独立站运营者设计** · 上传竞品广告素材 → AI深度拆解20个维度 → "
        "可迁移创意公式 + 品牌专属落地建议 → 一键导出Excel"
    )
    st.divider()

    # ══════════════════════════════════════
    # 侧边栏：API配置 + 品牌信息
    # ══════════════════════════════════════
    with st.sidebar:
        st.header("⚙️ API 配置")

        api_key = st.text_input(
            "Gemini API Key",
            type="password",
            placeholder="AIza...",
            help="访问 aistudio.google.com 免费注册获取，每天约1500次免费调用"
        )
        if not api_key:
            st.markdown("👉 [点此免费获取 Gemini API Key](https://aistudio.google.com)")

        model_name = st.selectbox(
            "分析模型",
            ["gemini-2.0-flash", "gemini-2.0-flash-lite", "gemini-2.5-pro-preview-03-25"],
            help="2.0 Flash：速度快成本低，推荐日常使用\n2.5 Pro：分析最深入，适合重要竞品研究"
        )

        st.divider()

        st.header("🏷️ 我的品牌信息")
        st.caption("填写越详细，AI给出的落地建议越精准。可不填，但会影响报告质量。")

        brand_name = st.text_input(
            "我的品牌名称",
            placeholder="填你自己的品牌名",
            help="你自己的品牌名，用于生成「针对XX品牌的落地建议」"
        )

        competitor_name = st.text_input(
            "被分析的竞品品牌",
            placeholder="填你上传素材所属的竞品品牌名",
            help="你上传的素材来自哪个竞品品牌？"
        )

        brand_product = st.text_area(
            "我的产品介绍",
            height=90,
            placeholder=(
                "填写你的产品类型、核心功效/特性、\n"
                "主要材质/成分、售价区间等\n"
                "（越具体，建议越精准）"
            ),
            help="产品类型、核心卖点、价位段、主要功效/特性"
        )

        brand_persona = st.text_area(
            "我的目标人群",
            height=90,
            placeholder=(
                "填写你的核心目标用户：\n"
                "年龄段、性别、生活状态\n"
                "他们最大的痛点或需求是什么？"
            ),
            help="年龄、性别、生活状态、核心痛点"
        )

        brand_usp = st.text_area(
            "我的核心差异化卖点",
            height=90,
            placeholder=(
                "你跟竞品最大的不同是什么？\n"
                "为什么用户应该选你而不是竞品？\n"
                "你最想让用户记住的一句话是？"
            ),
            help="你跟竞品最大的不同是什么？为什么用户选你不选竞品？"
        )

        st.divider()
        st.caption("🔒 API Key 仅在你的浏览器本地使用，不会被保存或传输到任何第三方")

    # ══════════════════════════════════════
    # 主区域：使用说明
    # ══════════════════════════════════════
    with st.expander("📖 使用步骤（点击展开）", expanded=False):
        st.markdown("""
**Step 1** → 在左侧填写你的 **Gemini API Key**（[免费获取](https://aistudio.google.com)）

**Step 2** → 在左侧填写你的**品牌信息**（品牌名、产品、人群、卖点）

**Step 3** → 打开 [Meta广告素材库](https://www.facebook.com/ads/library/)，找到竞品品牌，下载广告素材
- 图片：右键保存
- 视频：用录屏工具或下载插件保存为 MP4

**Step 4** → 点击「Browse files」**批量上传**素材

**Step 5** → 点击「开始分析」，等待 AI 自动拆解（每条约 10-20 秒）

**Step 6** → 查看分析表格和综合策略报告，**一键下载 Excel**

---
支持格式：图片（JPG、PNG、GIF）· 视频（MP4、MOV）

建议每批 **5-15条** 素材，超过 20 条建议分批上传
        """)

    # ══════════════════════════════════════
    # 素材上传
    # ══════════════════════════════════════
    st.header("📤 上传竞品广告素材")

    uploaded = st.file_uploader(
        "从Facebook Ads Library下载的广告素材（支持批量上传）",
        type=["jpg", "jpeg", "png", "gif", "mp4", "mov"],
        accept_multiple_files=True
    )

    if uploaded:
        st.subheader(f"✅ 已上传 {len(uploaded)} 个素材，预览如下")
        st.caption("AI 将直接分析素材的视觉内容，无需填写文案。")

        # 每行3列的预览网格
        cols_per_row = 3
        for i in range(0, len(uploaded), cols_per_row):
            cols = st.columns(cols_per_row)
            for j, col in enumerate(cols):
                idx = i + j
                if idx < len(uploaded):
                    f = uploaded[idx]
                    ext = Path(f.name).suffix.lower()
                    with col:
                        f.seek(0)
                        if ext in {'.jpg', '.jpeg', '.png', '.gif'}:
                            st.image(f, caption=f"AD-{str(idx+1).zfill(3)}",
                                     use_container_width=True)
                        else:
                            st.video(f)
                            st.caption(f"AD-{str(idx+1).zfill(3)}: {f.name}")

        st.divider()

        # ── 分析按钮 ──
        if not api_key:
            st.warning("⚠️ 请先在左侧输入 Gemini API Key 才能开始分析")

        _, btn_col, _ = st.columns([1, 2, 1])
        with btn_col:
            go = st.button(
                f"🚀 开始分析 {len(uploaded)} 条素材",
                type="primary",
                disabled=not api_key,
                use_container_width=True
            )

        # ══════════════════════════════════════
        # 执行分析
        # ══════════════════════════════════════
        if go and api_key:
            client = genai.Client(api_key=api_key)

            results = []
            progress_bar = st.progress(0, "准备开始分析...")
            status_msg = st.empty()

            for i, f in enumerate(uploaded):
                ad_id = f"AD-{str(i+1).zfill(3)}"
                status_msg.info(f"🔍 正在分析 **{ad_id}**：{f.name}  （{i+1} / {len(uploaded)}）")
                progress_bar.progress(i / len(uploaded))

                # 重置文件指针再读取（避免preview已读过）
                f.seek(0)
                file_bytes = f.read()
                ext = Path(f.name).suffix.lower()

                result = analyze_single_ad(client, model_name, ad_id, file_bytes, ext)
                result["_filename"] = f.name
                results.append(result)

                time.sleep(0.8)  # 避免触发API速率限制

            progress_bar.progress(1.0)
            status_msg.success(f"✅ {len(results)} 条素材全部分析完成！")

            # ── 显示明细表格 ──
            st.header("📊 分析结果明细")

            rows = []
            for r in results:
                row = {col: r.get(col, "—") for col in COLUMNS}
                rows.append(row)
            df = pd.DataFrame(rows)

            st.dataframe(df, use_container_width=True, height=420)

            # ── 生成综合策略报告 ──
            st.header("💡 综合策略报告")

            with st.spinner("🧠 AI正在生成竞品洞察和品牌落地建议，请稍候（约30-60秒）..."):
                # 准备送给AI的分析结果（去掉内部字段，控制长度）
                clean_results = [
                    {k: v for k, v in r.items() if not k.startswith("_")}
                    for r in results
                ]
                analysis_json = json.dumps(clean_results, ensure_ascii=False, indent=2)
                if len(analysis_json) > 12000:
                    analysis_json = analysis_json[:12000] + "\n...(内容较多，已节选)"

                sp = SUMMARY_PROMPT.format(
                    n=len(results),
                    analysis_json=analysis_json,
                    brand_name=brand_name or "我的品牌",
                    competitor_name=competitor_name or "参考竞品",
                    brand_product=brand_product or "未填写",
                    brand_persona=brand_persona or "未填写",
                    brand_usp=brand_usp or "未填写"
                )

                try:
                    summary_resp = client.models.generate_content(
                        model=model_name,
                        contents=sp
                    )
                    summary_text = summary_resp.text
                    st.markdown(summary_text)
                except Exception as e:
                    summary_text = f"综合报告生成遇到问题：{e}"
                    st.error(summary_text)

            # ── 导出Excel ──
            st.divider()
            st.header("📥 导出完整报告")

            _, dl_col, _ = st.columns([1, 2, 1])
            with dl_col:
                timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M')
                fname = f"META广告竞品分析_{competitor_name or 'report'}_{timestamp}.xlsx"
                st.download_button(
                    label="⬇️ 下载 Excel 完整报告（明细 + 策略报告）",
                    data=build_excel(df, summary_text),
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )

            st.success("🎉 分析完成！Excel报告包含两个Sheet：「素材分析明细」和「综合策略报告」")

    else:
        # 空状态引导
        st.info("👆 点击上方「Browse files」上传从 Meta 广告素材库下载的竞品广告素材")

        with st.expander("👀 示例输出效果预览"):
            sample = {
                "素材编号": ["AD-001", "AD-002", "AD-003"],
                "素材类型": ["视频", "图片", "视频"],
                "素材形式": ["UGC真实用户", "专业棚拍", "开箱测评"],
                "广告感判断": ["弱广告感（原生内容感）", "中等广告感", "弱广告感（原生内容感）"],
                "素材时长": [28, 0, 45],
                "Hook类型": ["痛点刺激型", "数据权威型", "好奇悬念型"],
                "核心信息主张": [
                    "解决长期困扰的敏感肌问题，只需30天",
                    "临床证明：98%用户4周内看到明显改善",
                    "皮肤科医生都在用的这款精华，你知道吗？"
                ],
                "适合投放阶段": ["TOF冷启动", "MOF考虑层", "TOF冷启动"]
            }
            st.dataframe(pd.DataFrame(sample), use_container_width=True)


if __name__ == "__main__":
    main()
